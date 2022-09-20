import csv
import os

import openpyxl

os.chdir(r'C:\Users\jegat\Desktop\py_work\work\project\converted')
os.makedirs('csv', exist_ok=True)

for file in os.listdir("."):
    # Skip non-xlsx files, load the workbook object.
    if not file.endswith(".xlsx"):
        continue
    wb = openpyxl.load_workbook(file)

    # Loop through every sheet in the workbook.
    for sheetNames in wb.get_sheet_names():
        # Create the CSV filename from the Excel filename and sheet title.
        sheet = wb.get_sheet_by_name(sheetNames)
    
        with open(os.path.join('csv', f'{file[:-5]}-{sheet.title}.csv'), "w", newline="") as csvFileObj:
            # Create the csv.writer object for this CSV file.
            writerObj = csv.writer(csvFileObj)
            # We need nested list to properly append data into csv files, because that's how csv files can be writtern.
            # This completeData list will hold the nested list.
            completeData = []
    
            # Loop through every row in the sheet.
            for rowNum in range(1, int(sheet.max_row) + 1):
                # This list will hold a single row of data in it.                 
                rowData = []
                # Loop through each cell in the row.
                for colNum in range(1, int(sheet.max_column) + 1):
                    # Append each cell's data.
                    rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                # Append each row of list into completeData list.
                completeData.append(rowData)

            # Write the rowData list to the CSC file.
            for data in completeData:
                writerObj.writerow(data)
